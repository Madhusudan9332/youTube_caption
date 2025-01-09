import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import time
from openpyxl import Workbook, load_workbook
from datetime import datetime
import csv
import os
import threading
import json

# Global variables for control
is_paused = False
is_stopped = False
input_file = None
append_file = None
current_index = 0
print_transcript_data = False
json_file = True

# Function to process videos
def process_videos(input_file, input_file_type, output_file, output_file_type):
    global is_paused, is_stopped, current_index, append_file, print_transcript_data, json_file

    if append_file:
        output_file = append_file

    # Define the required headers
    required_headers = ["URL", "Title", "Caption Available", "Time", "Text", "Data"]

    # Read input file
    if input_file_type == "xlsx":
        workbook = load_workbook(input_file)
        sheet = workbook.active
        video_urls = [row[0].value for row in sheet.iter_rows(min_row=2, max_col=1) if row[0].value]
    elif input_file_type == "csv":
        with open(input_file, "r", encoding="utf-8") as f:
            reader = csv.reader(f)
            next(reader)  # Skip header row
            video_urls = [row[0] for row in reader if row]

    # Read existing URLs from append file (if available)
    existing_urls = []
    if append_file:
        if append_file.endswith(".xlsx"):
            try:
                append_workbook = load_workbook(append_file)
                append_sheet = append_workbook.active

                # Check for headers
                append_headers = [cell.value for cell in append_sheet[1]]
                if append_headers != required_headers:
                    # Add missing headers
                    append_sheet.insert_rows(1)
                    for idx, header in enumerate(required_headers, start=1):
                        append_sheet.cell(row=1, column=idx, value=header)
                    append_workbook.save(append_file)

                # Extract existing URLs
                existing_urls = [row[0].value for row in append_sheet.iter_rows(min_row=2, max_col=1) if row[0].value]
            except Exception as e:
                print(f"Error reading append file (Excel): {e}")
        elif append_file.endswith(".csv"):
            try:
                with open(append_file, "r", encoding="utf-8") as f:
                    reader = csv.reader(f)
                    append_headers = next(reader, None)

                    # Check for headers
                    if append_headers != required_headers:
                        with open(append_file, "w", newline="", encoding="utf-8") as fw:
                            writer = csv.writer(fw)
                            writer.writerow(required_headers)  # Write headers
                            fw.writelines(f)  # Append existing content

                    # Extract existing URLs
                    existing_urls = [row[0] for row in reader if row]
            except Exception as e:
                print(f"Error reading append file (CSV): {e}")

    # Filter out URLs that are already in the append file
    video_urls = [url for url in video_urls if url not in existing_urls]
    print(f"Filtered URLs to process: {video_urls}")

    # Prepare the Selenium WebDriver
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service)

    try:
        # Initialize output data structure
        output_data = [["URL", "Title", "Caption Available", "Time", "Text", "Data"]]
        output_json_data = []

        # Iterate over each video URL
        while current_index < len(video_urls):
            if is_stopped:
                break
            if is_paused:
                time.sleep(1)
                continue

            video_url = video_urls[current_index]
            current_index += 1
            
            try:
                driver.get(video_url)
                driver.implicitly_wait(10)

                # Extract video title
                try:
                    title_element = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, ".ytd-watch-metadata #title"))
                    )
                    video_title = title_element.text.strip()
                    print(f"Video Title: {video_title}")
                except Exception as e:
                    print(f"Error extracting title: {e}")
                    video_title = "default_title"

                # Check caption availability using the 'fill-opacity' attribute
                try:
                    captions_button = driver.find_element(By.CSS_SELECTOR, ".ytp-subtitles-button svg")
                    fill_opacity = captions_button.get_attribute("fill-opacity")
                    Caption = False if not fill_opacity or float(fill_opacity) < 1 else True
                    print(f"Caption availability: {Caption}")
                except Exception as e:
                    Caption = False
                    print(f"Error checking captions availability: {e}")
                    continue
                
                # Locate and click 'Show transcript' button if captions are available
                segment_data = []
                segment_time_ = []
                segment_text_ = []
                # Step 2: Collect time and text data from the transcript
                transcript_data = []
                if Caption:
                    # mute video
                    try:
                        mute_button = WebDriverWait(driver, 10).until(
                            EC.element_to_be_clickable((By.CSS_SELECTOR, "button.ytp-mute-button"))
                        )
                        mute_button.click()
                        print("Clicked on the mute/unmute button successfully!")
                    except Exception as e:
                        print(f"An error occurred: {e}")
                    # Locate the 'expand' button and click it
                    try:
                        expand_button = driver.find_element(By.CSS_SELECTOR, "tp-yt-paper-button#expand")
                        expand_button.click()
                        time.sleep(2)
                    except Exception as e:
                        print(f"Error clicking expand button: {e}")
                        continue
                    # Click to Show Transcript
                    try:
                        transcript_button = driver.find_element(By.XPATH, "//button[contains(@aria-label, 'Show transcript')]")
                        transcript_button.click()
                        time.sleep(3)

                        # Wait for transcript segments to load
                        WebDriverWait(driver, 10).until(
                            EC.presence_of_all_elements_located((By.CSS_SELECTOR, "ytd-transcript-segment-renderer"))
                        )

                        # Collect transcript data
                        try:
                            segments = driver.find_elements(By.CSS_SELECTOR, "ytd-transcript-segment-renderer")
                            if segments:
                                print("Transcript segments found: collectig data...")
                            for segment in segments:
                                try:
                                    time_element = segment.find_element(By.CSS_SELECTOR, ".segment-timestamp")
                                    segment_time = time_element.text.strip()
                                    text_element = segment.find_element(By.CSS_SELECTOR, ".segment-text")
                                    segment_text = text_element.text.strip()
                                    segment_time_.append(segment_time)
                                    segment_text_.append(segment_text)
                                    segment_data.append([segment_time, segment_text])
                                    if json_file.get():                                        
                                        transcript_data.append({"time": segment_time, "text": segment_text})
                                    if print_transcript_data.get():
                                        print(f"Time: {segment_time} - Text: {segment_text}")
                                except Exception as e:
                                    print(f"Error extracting data: {e}")
                            print("Transcript data collected successfully!")
                        except Exception as e:
                            print(f"Error collecting transcript data: {e}")                        
                    except Exception as e:
                        print(f"Error clicking transcript button")

                # Flatten segment data to a string format
                segment_data_str = ", ".join([f"{item[0]}: {item[1]}" for item in segment_data])

                # Append the data
                output_data.append([
                    video_url,
                    video_title,
                    Caption,
                    segment_data_str,  # Flattened segment data
                    ", ".join(segment_time_),  # Times as a single string
                    " | ".join(segment_text_)  # Texts as a single string
                ])
                if json_file.get():
                    output_json_data.append({
                        "URL": video_url,
                        "Title": video_title,
                        "Caption Available": Caption,
                        "Data": transcript_data
                    })

            except Exception as e:
                print(f"Error processing URL {video_url}: {e}")

        # Write to the output file
        if json_file.get():
            with open(f"{output_file}.json", "w", encoding="utf-8") as json_file:
                json.dump(output_json_data, json_file, ensure_ascii=False, indent=4)
        print("Transcript data saved to 'transcript_data.json'.")
        if output_file_type == "xlsx":
            if append_file:
                workbook = load_workbook(output_file)
                sheet = workbook.active
            else:
                workbook = Workbook()
                sheet = workbook.active
            for row in output_data:
                sheet.append(row)
            workbook.save(output_file)
        elif output_file_type == "csv":
            with open(output_file, "a" if append_file else "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerows(output_data)

        # Notify the user
        messagebox.showinfo("Process Complete", f"Transcript saved to {output_file}")

    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")

    finally:
        driver.quit()

# Function to browse and select input file
def browse_input_file():
    input_file_type = input_type_var.get()
    filetypes = [("Excel Files", "*.xlsx")] if input_file_type == "xlsx" else [("CSV Files", "*.csv")]

    input_file_path = filedialog.askopenfilename(title="Select Input File", filetypes=filetypes)

    if input_file_path:
        global input_file
        input_file = input_file_path
        file_label.config(text=f"Selected File: {os.path.basename(input_file)}")

# Function to browse and select append file
def browse_append_file():
    global append_file
    output_file_type = output_type_var.get()
    filetypes = [("Excel Files", "*.xlsx")] if output_file_type == "xlsx" else [("CSV Files", "*.csv")]
    append_file_path = filedialog.askopenfilename(title="Select Append File", filetypes=filetypes)

    if append_file_path:
        global append_file
        append_file = append_file_path
        append_label.config(text=f"Append File: {os.path.basename(append_file)}")
      
# Clear append file 
def clear_append_file():
    global append_file
    append_file = None
    append_label.config(text="No append file selected")
  
# Start the process
def start_process():
    global is_paused, is_stopped, current_index
    is_paused = False
    is_stopped = False
    current_index = 0

    input_file_type = input_type_var.get()
    output_file_type = output_type_var.get()
    if not input_file or not input_file_type or not output_file_type:
        messagebox.showerror("Error", "Please select an input file and output format!")
        return

    # Generate output file name with timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = f"youtube_transcripts_{timestamp}.{output_file_type}"
    
    # Run the process in a separate thread
    threading.Thread(target=process_videos, args=(input_file, input_file_type, output_file,output_file_type)).start()

def toggle_pause_resume():
    global is_paused
    is_paused = not is_paused  # Toggle the pause state

    # Update the button text
    if is_paused:
        pause_button.config(text="Resume", bg="orange")
    else:
        pause_button.config(text="Pause", bg="yellow")

# Stop the process
def stop_process():
    global is_stopped
    is_stopped = True


# Create the GUI window
def main():
    global input_file, file_label, append_label, input_type_var, output_type_var, pause_button, print_transcript_data, json_file

    input_file = None

    root = tk.Tk()
    
    print_transcript_data = tk.BooleanVar(value=False)
    json_file = tk.BooleanVar(value=True)
    
    root.title("YouTube Transcript Downloader")
    root.geometry("500x500")

    tk.Label(root, text="Select Input File Type:", font=("Arial", 12)).pack(pady=10)
    input_type_var = tk.StringVar(value="xlsx")
    ttk.Combobox(root, textvariable=input_type_var, values=["xlsx", "csv"], state="readonly").pack(pady=5)

    tk.Label(root, text="Select Output File Format:", font=("Arial", 12)).pack(pady=10)
    output_type_var = tk.StringVar(value="xlsx")
    ttk.Combobox(root, textvariable=output_type_var, values=["xlsx", "csv"], state="readonly").pack(pady=5)
    
    checkbox_print = tk.Checkbutton(root, text="Print Transcript Data", variable=print_transcript_data)
    label_print = tk.Label(root, text=f"Checkbox Value: {print_transcript_data.get()}")
    checkbox_print.pack(pady=5)
    label_print.pack(pady=10)
    
    checkbox_json = tk.Checkbutton(root, text="Json File", variable=json_file)
    label_json = tk.Label(root, text=f"Checkbox Value: {json_file.get()}")
    checkbox_json.pack(pady=5)
    label_json.pack(pady=10)
    
    # Input and Append File Selection (Inline Buttons)
    file_frame = tk.Frame(root)
    file_frame.pack(pady=10)
    
    tk.Button(file_frame, text="Browse Input File", command=browse_input_file, font=("Arial", 12), bg="blue", fg="white").grid(row=0, column=0, padx=5)
    tk.Button(file_frame, text="Append Output File", command=browse_append_file, font=("Arial", 12), bg="purple", fg="white").grid(row=0, column=1, padx=5)
    tk.Button(file_frame, text="Clear Append File", command=clear_append_file, font=("Arial", 12), bg="red", fg="white").grid(row=0, column=2, padx=5)

    file_label = tk.Label(root, text="No file selected", font=("Arial", 10))
    file_label.pack(pady=5)

    append_label = tk.Label(root, text="No append file selected", font=("Arial", 10))
    append_label.pack(pady=5)
    
    # Buttons for Start, Pause, and Stop
    button_frame = tk.Frame(root)
    button_frame.pack(pady=10)

    # Start Button
    tk.Button(button_frame, text="Start Process", command=start_process, font=("Arial", 12), bg="green", fg="white").grid(row=0, column=0, padx=5)
    # Pause/Resume Button
    pause_button = tk.Button(button_frame, text="Pause", command=toggle_pause_resume, font=("Arial", 12), bg="yellow", fg="black")
    pause_button.grid(row=0, column=1, padx=5)
    # Stop Button
    tk.Button(button_frame, text="Stop", command=stop_process, font=("Arial", 12), bg="red", fg="white").grid(row=0, column=2, padx=5)


    root.mainloop()


if __name__ == "__main__":
    main()
